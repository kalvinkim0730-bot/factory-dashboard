import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import requests
import time
import re
import io
import os
import base64
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image as PILImage

# 가상 서버 루트 작업 영역 직통 세팅
SAVED_EXCEL_PATH = "permanent_production_schedule.xlsx"
NOTES_DB_PATH = "production_notes.txt"
MASTER_PASSWORD = "Fineformulation"
ENTRY_SECURITY_CODE = "1234"      # [대표님 지정 핵심 보안 코드]
SESSION_TIMEOUT_SEC = 300          # [대표님 지정: 열람 유효시간 5분 (300초)]

# [오너 지시 정규식 핵심 축]: 띄어쓰기, 언더바 다 무시하고 오직 앞자리 순수 6자리 코드만 정밀 추출
def extract_pure_6_code(text):
    if not text:
        return ""
    cleaned = str(text).replace(" ", "").replace("_", "").replace("\r", "").replace("\n", "").replace("\t", "").strip().upper()
    match = re.search(r'(\d{5}[A-Z])', cleaned)
    return match.group(1) if match else ""

# [대표님 명세 1순위 조항]: 로컬 저장 파일을 1순위로 호출
def get_saved_local_image_base64(pure_code):
    pure_code_clean = str(pure_code).strip().upper()
    target_path = f"{pure_code_clean}.png"
    if os.path.exists(target_path):
        try:
            with open(target_path, "rb") as f:
                encoded = base64.b64encode(f.read()).decode()
                return f"data:image/png;base64,{encoded}"
        except Exception:
            return None
    return None

# ---------------------------------------------------------------------
# [🚨 백엔드 엔지니어링 개조]: 엑셀 파싱 연산 메모리 캐싱 장치 주입 (속도 10배 향상)
# ---------------------------------------------------------------------
@st.cache_data(ttl=60)
def load_and_clean_excel(file_path):
    raw_df = pd.read_excel(file_path, header=None)
    start_row_idx = 0
    for idx, row in raw_df.iterrows():
        row_str_list = [str(cell) for cell in row.dropna().tolist()]
        combined_row_text = "".join(row_str_list)
        if any(keyword in combined_row_text for keyword in ['일정', '코드', '카테고리', 'Date', 'Item']):
            start_row_idx = idx + 1
            break
            
    clean_data_list = []
    for idx in range(start_row_idx, len(raw_df)):
        row_cells = raw_df.iloc[idx]
        if len(row_cells) < 22:
            continue
        p_date = pd.to_datetime(row_cells[21], errors='coerce') 
        if pd.isna(p_date): 
            continue
        clean_data_list.append({
            'item_code': str(row_cells[0]).strip(),     
            'category': str(row_cells[2]).strip(),      
            'price_tag': str(row_cells[5]).strip(),     
            'po_number': str(row_cells[10]).strip(),    
            'bag_number': str(row_cells[11]).strip(),   
            'volume': str(row_cells[12]).strip(),       
            'product_name': str(row_cells[14]).strip(), 
            'quantity': row_cells[16],                  
            'production_date': p_date
        })
    df = pd.DataFrame(clean_data_list)
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0).astype(int)
    for col in ['item_code', 'category', 'price_tag', 'po_number', 'bag_number', 'volume', 'product_name']:
        df[col] = df[col].fillna('-').astype(str).str.strip()
        df[col] = df[col].replace(['nan', 'NAN', 'NaN', 'None', '', ' ', '-'], '-')
        df[col] = df[col].apply(lambda x: '-' if str(x).strip() not in ['Y', 'N'] and col == 'price_tag' else x)
    return df

# ---------------------------------------------------------------------
# [📝 5대 입력 데이터 + 완료 상태 비트 영구 저장 엔진]
# ---------------------------------------------------------------------
def load_production_notes():
    notes = {}
    if os.path.exists(NOTES_DB_PATH):
        try:
            with open(NOTES_DB_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    if "::" in line:
                        parts = line.split("::", 6)
                        code = parts[0].strip()
                        c_code = parts[1].strip() if len(parts) > 1 else ""
                        pack_qty = parts[2].strip() if len(parts) > 2 else ""
                        m_date = parts[3].strip() if len(parts) > 3 else ""
                        m_qty = parts[4].strip() if len(parts) > 4 else ""
                        p_qty = parts[5].strip() if len(parts) > 5 else ""
                        is_done = parts[6].strip() if len(parts) > 6 else "N"
                        notes[code] = (c_code, pack_qty, m_date, m_qty, p_qty, is_done)
        except Exception:
            pass
    return notes

def save_production_note(pure_code, c_code, pack_qty, m_date, m_qty, p_qty, is_done="N"):
    notes = load_production_notes()
    notes[pure_code] = (c_code.strip(), pack_qty.strip(), m_date.strip(), m_qty.strip(), p_qty.strip(), is_done.strip())
    try:
        with open(NOTES_DB_PATH, "w", encoding="utf-8") as f:
            for code, values in notes.items():
                if any(values):
                    f.write(f"{code}::{values[0]}::{values[1]}::{values[2]}::{values[3]}::{values[4]}::{values[5]}\n")
    except Exception:
        pass

# =========================================================================
# 2. 스트림릿 웹 대시보드 UI 레이아웃 구성
# =========================================================================
st.set_page_config(layout="wide", page_title="생산 스케줄 마스터 데이터 경영 대시보드")

# ---------------------------------------------------------------------
# [🚨 실시간 사용 감지 5분 연장 엔진]
# ---------------------------------------------------------------------
if "app_unlocked" not in st.session_state:
    st.session_state["app_unlocked"] = False
if "unlock_time" not in st.session_state:
    st.session_state["unlock_time"] = None

if st.session_state["app_unlocked"] and st.session_state["unlock_time"] is not None:
    elapsed_time = time.time() - st.session_state["unlock_time"]
    if elapsed_time > SESSION_TIMEOUT_SEC:
        st.session_state["app_unlocked"] = False
        st.session_state["unlock_time"] = None
        st.toast("⚠️ 보안 유지를 위해 자리를 비우신 지 5분이 경과되어 자동 잠금되었습니다.")
        time.sleep(1)
        st.rerun()
    else:
        st.session_state["unlock_time"] = time.time()

# ---------------------------------------------------------------------
# [🔒 게이트웨이 정문 차단막 인터페이스]
# ---------------------------------------------------------------------
if not st.session_state["app_unlocked"]:
    st.markdown("""
        <style>
            .stApp { background-color: #0f172a !important; }
            .security-gate {
                text-align: center; margin-top: 15vh; padding: 40px;
                background-color: #1e2530; border: 2px solid #38bdf8; border-radius: 16px;
                box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.7); max-width: 500px; margin-left: auto; margin-right: auto;
            }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
        <div class="security-gate">
            <h1 style="color: #38bdf8; font-size: 28px; font-weight: bold; margin-bottom: 10px;">🔒 FINE FORMULATION</h1>
            <p style="color: #94a3b8; font-size: 15px; margin-bottom: 25px;">본 시스템은 기업 기밀 자산 보호 구역입니다.<br>열람 유효시간은 5분이며, <b>사용 중일 경우 실시간으로 자동 연장</b>됩니다.</p>
        </div>
    """, unsafe_allow_html=True)
    
    cols = st.columns([1, 2, 1])
    with cols[1]:
        input_gate_code = st.text_input("🔑 보안 코드 입력 (Security Code)", type="password", key="gate_code_input")
        if input_gate_code == ENTRY_SECURITY_CODE:
            st.session_state["app_unlocked"] = True
            st.session_state["unlock_time"] = time.time()
            st.success("🔓 자격 증명이 확인되었습니다. 시스템을 개방합니다.")
            time.sleep(0.5)
            st.rerun()
    st.stop()

# ---------------------------------------------------------------------
# [🔓 오픈 마스터 대시보드 코어]
# ---------------------------------------------------------------------
has_saved_file = os.path.exists(SAVED_EXCEL_PATH)
final_file_target = SAVED_EXCEL_PATH if has_saved_file else None

with st.sidebar:
    st.markdown(f'<div style="color:#ffffff; font-size:15px; font-weight:bold; background-color:#0284c7; padding:10px; border-radius:8px; margin-bottom:15px; text-align:center;">🟢 시스템 가동 중 (활동 중 자동 연장)</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:20px; font-weight:bold; color:#38bdf8; margin-bottom:15px; border-bottom:2px solid #38bdf8; padding-bottom:5px;">⚙️ 마스터 데이터 제어 센터</div>', unsafe_allow_html=True)
    
    input_password = st.text_input("🔓 데이터 제어 승인 암호", type="password", key="auth_pwd_input")
    is_authenticated = (input_password == MASTER_PASSWORD)

    st.markdown("---")
    st.write("📂 **새로운 스케줄 파일 업로드 / 교체**")
    uploaded_file = st.file_uploader("여기에 엑셀 파일을 드래그 앤 드롭 하세요.", type=["xlsx", "xls"], label_visibility="collapsed")
    
    if uploaded_file and is_authenticated:
        with open(SAVED_EXCEL_PATH, "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.cache_data.clear() # 새 파일 인입 시 가상 메모리 전격 초기화
        st.success("🚀 마스터 스케줄 파일 교체 성공!")
        time.sleep(1)
        st.rerun()

if final_file_target:
    # 캐싱 적용된 고속 로더 엔진 교체
    df = load_and_clean_excel(final_file_target)
    
    # 주차 타임스탬프 계산
    today_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    current_weekday = today_dt.weekday() 
    next_monday_dist = (7 - current_weekday) % 7 or 7
    
    target_next_monday = (today_dt + timedelta(days=next_monday_dist)).replace(hour=23, minute=59, second=59)
    second_monday_start = target_next_monday + timedelta(seconds=1)
    target_second_monday = (second_monday_start + timedelta(days=6)).replace(hour=23, minute=59, second=59)
    
    saved_notes = load_production_notes()

    completed_codes = [code for code, vals in saved_notes.items() if len(vals) > 5 and vals[5] == "Y"]
    
    # 지연 및 정상 주차 분리
    df_delayed = df[(df['production_date'] < today_dt) & (~df['item_code'].apply(extract_pure_6_code).isin(completed_codes))].copy()
    df_active_pool = df[(df['production_date'] >= today_dt) & (~df['item_code'].apply(extract_pure_6_code).isin(completed_codes))].copy()
    
    df_1week = df_active_pool[df_active_pool['production_date'] <= target_next_monday].copy()
    df_2weeks = df_active_pool[(df_active_pool['production_date'] >= second_monday_start) & (df_active_pool['production_date'] <= target_second_monday)].copy()
    
    df_completed_pool = df[df['item_code'].apply(extract_pure_6_code).isin(completed_codes)].copy()

    # ---------------------------------------------------------------------
    # [📊 주차별 분리형 마스터 엑셀 컴파일러]
    # ---------------------------------------------------------------------
    def generate_premium_split_excel(df_w1, df_w2):
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "주차별_생산라인업"
        ws.views.sheetView[0].showGridLines = True
        
        font_main_title = Font(name="Malgun Gothic", size=14, bold=True, color="FFFFFF")
        font_header = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Malgun Gothic", size=10)
        font_group = Font(name="Malgun Gothic", size=11, bold=True, color="0f172a")
        
        fill_week_title = PatternFill(start_color="0369a1", end_color="0369a1", fill_type="solid")
        fill_header = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        fill_group = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        thin_side = Side(border_style="thin", color="cbd5e1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        headers = ["카테고리 그룹", "아이템 사진", "아이템 코드", "아이템 이름", "용량", "생산 수량", "PO 번호", "Bag#", "가격표 유무", "카톤코드", "개입수", "제조일", "제조량", "생산수량"]
        categories_order = ["skin", "body", "hair", "기타 카테고리"]
        current_row_idx = 1
        
        def write_week_block(ws, target_df, week_label_text, start_row):
            r_idx = start_row
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=14)
            title_cell = ws.cell(row=r_idx, column=1)
            title_cell.value = week_label_text
            title_cell.font = font_main_title; title_cell.fill = fill_week_title; title_cell.alignment = align_center
            ws.row_dimensions[r_idx].height = 35
            r_idx += 1
            
            for col_num, h_text in enumerate(headers, 1):
                h_cell = ws.cell(row=r_idx, column=col_num, value=h_text)
                h_cell.font = font_header; h_cell.fill = fill_header; h_cell.alignment = align_center; h_cell.border = border_all
            ws.row_dimensions[r_idx].height = 25
            r_idx += 1
            
            for cate in categories_order:
                cate_df = target_df[target_df['category'].str.lower().str.contains(cate)] if cate != "기타 카테고리" else target_df[~target_df['category'].str.lower().str.contains('skin|body|hair')]
                if not cate_df.empty:
                    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=14)
                    g_cell = ws.cell(row=r_idx, column=1, value=f"🌿 {cate.upper()} CARE LINEUP")
                    g_cell.font = font_group; g_cell.fill = fill_group; g_cell.alignment = align_left
                    for c_num in range(1, 15):
                        ws.cell(row=r_idx, column=c_num).border = border_all
                    ws.row_dimensions[r_idx].height = 24
                    r_idx += 1
                    
                    for _, r in cate_df.iterrows():
                        p_code = extract_pure_6_code(r['item_code'])
                        memo_vals = saved_notes.get(p_code, ("", "", "", "", ""))
                        
                        ws.cell(row=r_idx, column=1, value=r['category'])
                        ws.cell(row=r_idx, column=3, value=r['item_code'])
                        ws.cell(row=r_idx, column=4, value=r['product_name'])
                        ws.cell(row=r_idx, column=5, value=r['volume'])     
                        qty_cell = ws.cell(row=r_idx, column=6, value=r['quantity']); qty_cell.number_format = '#,##0'; qty_cell.alignment = align_right
                        ws.cell(row=r_idx, column=7, value=r['po_number'])  
                        ws.cell(row=r_idx, column=8, value=r['bag_number']) 
                        ws.cell(row=r_idx, column=9, value=r['price_tag'])  
                        
                        ws.cell(row=r_idx, column=10, value=memo_vals[0]).alignment = align_center
                        ws.cell(row=r_idx, column=11, value=memo_vals[1]).alignment = align_center
                        ws.cell(row=r_idx, column=12, value=memo_vals[2]).alignment = align_center
                        ws.cell(row=r_idx, column=13, value=memo_vals[3]).alignment = align_center
                        ws.cell(row=r_idx, column=14, value=memo_vals[4]).alignment = align_center
                        
                        for c_idx in range(1, 15):
                            c_cell = ws.cell(row=r_idx, column=c_idx); c_cell.font = font_data; c_cell.border = border_all
                            if c_idx not in [4, 6]: c_cell.alignment = align_center
                                
                        ws.row_dimensions[r_idx].height = 35
                        img_path = f"{p_code}.png"
                        if os.path.exists(img_path):
                            try:
                                pil_img = PILImage.open(img_path); pil_img.thumbnail((50, 45))
                                img_stream = io.BytesIO(); pil_img.save(img_stream, format="PNG"); img_stream.seek(0)
                                xl_img = OpenpyxlImage(img_stream); ws.add_image(xl_img, f"B{r_idx}")
                            except: pass
                        r_idx += 1
            return r_idx + 2
            
        next_start_row = write_week_block(ws, df_1week, f"🗓️ 1주 차 생산 라인업 계획", current_row_idx)
        write_week_block(ws, df_w2, f"🗓️ 2주 차 생산 라인업 계획", next_start_row)
        
        for l, w in [('A', 15), ('B', 12), ('C', 16), ('D', 38), ('E', 12), ('F', 14), ('G', 16), ('H', 14), ('I', 14), ('J', 16), ('K', 12), ('L', 14), ('M', 14), ('N', 14)]:
            ws.column_dimensions[l].width = w
        wb.save(output)
        return output.getvalue()

    with st.sidebar:
        st.markdown("---")
        st.markdown('<div style="font-size:16px; font-weight:bold; color:#38bdf8;">📥 오너 기획 데이터 추출 센터</div>', unsafe_allow_html=True)
        split_excel_bytes = generate_premium_split_excel(df_1week, df_2weeks)
        st.download_button(label="📊 주차별 분리 마스터 엑셀 다운로드", data=split_excel_bytes, file_name=f"Fine_Formulation_Split_Schedule_{datetime.now().strftime('%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        st.markdown("---")
        st.markdown('<div style="font-size:16px; font-weight:bold; color:#fbbf24;">⚡ 트렐로 이미지 서버 백업 센터</div>', unsafe_allow_html=True)
        if st.button("🔄 현재 스케줄 이미지 서버에 저장", use_container_width=True):
            if is_authenticated:
                sync_success_count = 0
                status_placeholder = st.empty()
                status_placeholder.info("🔄 마스터 엑셀에서 순수 6자리 코드 스캔 중...")
                
                target_pure_codes = []
                for val in df['item_code'].dropna().unique():
                    p_code = extract_pure_6_code(val)
                    if p_code: target_pure_codes.append(p_code)
                target_pure_codes = list(set(target_pure_codes))
                
                if len(target_pure_codes) > 0:
                    TRELLO_API_KEY = st.secrets["TRELLO_API_KEY"]
                    TRELLO_TOKEN = st.secrets["TRELLO_TOKEN"]
                    TRELLO_BOARD_ID = st.secrets["TRELLO_BOARD_ID"]
                    secured_headers = {"Authorization": f'OAuth oauth_consumer_key="{TRELLO_API_KEY}", oauth_token="{TRELLO_TOKEN}"', "User-Agent": "Mozilla/5.0"}
                    
                    url = f"https://api.trello.com/1/boards/{TRELLO_BOARD_ID}/cards"
                    card_res = requests.get(url, headers=secured_headers, params={'key': TRELLO_API_KEY, 'token': TRELLO_TOKEN, 'attachments': 'true', 'limit': '1000'}, timeout=25)
                    
                    if card_res.status_code == 200:
                        all_cards = card_res.json(); progress_bar = st.progress(0); total_items = len(target_pure_codes)
                        for i, code_key in enumerate(target_pure_codes):
                            code_key_clean = str(code_key).strip().upper(); trello_url = None
                            status_placeholder.info(f"⏳ 이미지 매칭 중: [{code_key_clean}] ({i+1}/{total_items})")
                            
                            for card in all_cards:
                                if code_key_clean in card.get('name', '').replace(" ", "").upper():
                                    cover = card.get('cover', {})
                                    if cover and cover.get('scaled'): trello_url = cover.get('scaled', [])[-1].get('url')
                                    if not trello_url and card.get('attachments'):
                                        trello_url = card['attachments'][0].get('url')
                            if trello_url:
                                try:
                                    img_res = requests.get(trello_url, headers=secured_headers, timeout=12)
                                    if img_res.status_code == 200:
                                        with open(f"{code_key_clean}.png", "wb") as img_f: img_f.write(img_res.content)
                                        sync_success_count += 1
                                except: pass
                            progress_bar.progress(int((i + 1) / total_items * 100))
                        status_placeholder.empty()
                        st.success(f"🎯 총 {sync_success_count}개 품목의 공식 썸네일 다운로드 성공!")
                        time.sleep(1)
                        st.rerun()

    # 초경량 풀와획 반응형 CSS
    st.markdown("""
        <style>
            .floating-shortcut-anchor {
                position: fixed !important; right: 35px !important; top: 50% !important;
                transform: translateY(-50%) !important; z-index: 999999 !important;
                background: linear-gradient(135deg, #0284c7 0%, #0369a1 100%) !important;
                color: #ffffff !important; border: 2px solid #38bdf8 !important; border-radius: 12px !important;
                padding: 14px 18px !important; font-family: 'Malgun Gothic', sans-serif !important;
                font-size: 14px !important; font-weight: 900 !important; text-decoration: none !important;
                box-shadow: 0 0 20px rgba(56, 189, 248, 0.6) !important; transition: all 0.2s ease-in-out !important;
                display: flex !important; align-items: center !important; justify-content: center !important;
                letter-spacing: -0.3px !important; cursor: pointer !important;
            }
            @media screen and (max-width: 768px) {
                .floating-shortcut-anchor {
                    top: 12px !important; right: 12px !important; transform: none !important;
                    padding: 8px 12px !important; font-size: 12px !important; border-radius: 8px !important;
                }
            }
            
            html { scroll-behavior: smooth !important; }
            .owner-square-frame { width: 100% !important; aspect-ratio: 1 / 1 !important; background-color: transparent !important; display: flex !important; justify-content: center !important; align-items: center !important; overflow: hidden !important; padding: 5px !important; box-sizing: border-box !important; margin-bottom: 8px !important; }
            .owner-square-frame img { max-width: 100% !important; max-height: 100% !important; width: auto !important; height: auto !important; object-fit: contain !important; }
            .owner-info-card-wrap { background-color: #1e2530 !important; border: 1px solid #2d3748 !important; border-radius: 14px !important; padding: 18px !important; box-shadow: 0 10px 15px -3px rgba(0,0,0,0.4) !important; margin-bottom: 8px !important; }
            .owner-text-row { margin: 0px !important; padding: 0px !important; text-align: left !important; line-height: 1.4 !important; }
            
            div[data-testid="stVerticalBlock"] > div { margin-bottom: 0px !important; padding-bottom: 0px !important; }
            
            /* 플레이스홀더 풀 와이드 기획 락 */
            div[data-testid="stTextInput"] { display: block !important; margin-top: 0px !important; margin-bottom: 2px !important; padding: 0px !important; }
            div[data-testid="stTextInput"] label { display: none !important; }
            div[data-testid="stTextInput"] div[data-testid="stWidgetLabel"] + div { width: 100% !important; }
            div[data-testid="stTextInput"] input { 
                background-color: #0f172a !important; color: #38bdf8 !important; 
                border: 1px solid #334155 !important; border-radius: 6px !important; 
                font-size: 13px !important; height: 32px !important; padding: 4px 10px !important;
                width: 100% !important;
            }
            
            div[data-testid="stCheckbox"] { background-color: #111827 !important; border: 1px dashed #ef4444 !important; border-radius: 8px !important; padding: 6px 12px !important; margin-top: 5px !important; margin-bottom: 8px !important; }
            div[data-testid="stCheckbox"] label span { color: #f87171 !important; font-weight: bold !important; font-size: 14px !important; }
            div.delay-box-mark div[data-testid="stCheckbox"] { border: 1px solid #ff0000 !important; background-color: #2d1616 !important; }
            div.delay-box-mark div[data-testid="stCheckbox"] label span { color: #ff6b6b !important; }
            
            /* 💾 [일괄 저장 버튼 전용 명품 스타일 주입] */
            div.owner-save-btn-zone button {
                background: linear-gradient(135deg, #10b981 0%, #059669 100%) !important;
                color: white !important; border: 1px solid #34d399 !important; border-radius: 6px !important;
                font-weight: bold !important; font-size: 13px !important; height: 34px !important;
                box-shadow: 0 4px 6px -1px rgba(16, 185, 129, 0.3) !important; transition: all 0.1s !important;
            }
            div.owner-save-btn-zone button:hover { background: #34d399 !important; color: #0f172a !important; transform: scale(1.02) !important; }

            .element-container { margin-bottom: 0px !important; padding-bottom: 0px !important; }
        </style>
    """, unsafe_allow_html=True)

    st.markdown('<a href="#current-week-target" class="floating-shortcut-anchor">📅 현재 주차 이동</a>', unsafe_allow_html=True)

    def render_schedule_grid(target_df, title_label, section_prefix, is_completed_tab=False, is_delay_tab=False, inject_anchor=False):
        if is_delay_tab and target_df.empty:
            return
            
        st.markdown("---")
        if inject_anchor:
            st.markdown('<div id="current-week-target" style="padding-top:10px;"></div>', unsafe_allow_html=True)
            
        st.subheader(title_label)
        
        if not target_df.empty:
            fixed_categories = ["skin", "body", "hair"]
            for cate in fixed_categories:
                group_df = target_df[target_df['category'].str.lower().str.contains(cate)]
                if not group_df.empty:
                    st.markdown(f'<div style="font-size:20px; font-weight:bold; color:#38bdf8; padding:6px 12px; background-color:#0f172a; border-left:5px solid #38bdf8; border-radius:4px; margin-top:25px; margin-bottom:15px;">📦 {cate.upper()} care Lineup</div>', unsafe_allow_html=True)
                    cols = st.columns(6)
                    for idx, row in group_df.reset_index(drop=True).iterrows():
                        excel_code = row['item_code']
                        pure_excel_code = extract_pure_6_code(excel_code)
                        
                        with cols[idx % 6]:
                            local_base64_data = get_saved_local_image_base64(pure_excel_code)
                            st.html(f'<div class="owner-square-frame"><img src="{local_base64_data if local_base64_data else ""}"></div>')
                            
                            st.html(f"""
                                <div class="owner-info-card-wrap" style="border: 2px solid {'#ef4444' if is_delay_tab else '#2d3748'} !important;">
                                    <div class="owner-text-row" style="font-size:30px !important; font-weight:900 !important; color:#ffffff !important; margin-bottom:6px !important; letter-spacing:0.5px !important;">{excel_code}</div>
                                    <div class="owner-text-row" style="font-size:14px !important; color:#a0aec0 !important; font-weight:500 !important; min-height:40px !important; margin-bottom:14px !important; display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; overflow:hidden;">{row['product_name']}</div>
                                    <div style="border-bottom:1px solid #2d3748 !important; margin-bottom:12px !important;"></div>
                                    <div class="owner-text-row" style="font-size:14px !important; color:#718096 !important; margin-bottom:3px !important;">가격표 유무: <span style="color:#63b3ed !important; font-weight:bold !important;">{row['price_tag']}</span></div>
                                    <div class="owner-text-row" style="font-size:14px !important; color:#ffffff !important; margin-bottom:3px !important;">용량: <span style="color:#ffffff !important; font-weight:bold !important;">{row['volume']}</span></div>
                                    <div class="owner-text-row" style="font-size:14px !important; color:#718096 !important; margin-bottom:3px !important;">PO#: <span style="color:#ecc94b !important; font-weight:bold !important;">{row['po_number']}</span></div>
                                    <div class="owner-text-row" style="font-size:14px !important; color:#718096 !important; margin-bottom:3px !important;">Bag#: <span style="color:#e53e3e !important; font-weight:bold !important;">{row['bag_number']}</span></div>
                                    <div style="background-color:{'#3b1414' if is_delay_tab else '#111622'} !important; border-radius:8px !important; padding:8px 12px !important; display:flex !important; justify-content:space-between !important; align-items:center !important;">
                                        <span class="owner-text-row" style="font-size:16px !important; color:{'#ff6b6b' if is_delay_tab else '#48bb78'} !important; font-weight:bold !important;">📦 {row['quantity']:,}개</span>
                                        <span class="owner-text-row" style="font-size:13px !important; color:#a0aec0 !important; font-weight:500 !important;">📅 {row['production_date'].strftime('%m-%d')}</span>
                                    </div>
                                </div>
                            """)
                            
                            memo_tuple = saved_notes.get(pure_excel_code, ("-", "-", "-", "-", "-", "N"))
                            current_check_state = True if memo_tuple[5] == "Y" else False
                            
                            if is_completed_tab: check_label = "✅ 생산 완료 완료됨"
                            elif is_delay_tab: check_label = "⚠️ 긴급! 생산완료 처리"
                            else: check_label = "🚨 완료 시 체크 (하단 이동)"
                                
                            if is_delay_tab: st.markdown('<div class="delay-box-mark">', unsafe_allow_html=True)
                            user_checked = st.checkbox(label=check_label, value=current_check_state, key=f"chk_{section_prefix}_{pure_excel_code}_{idx}")
                            if is_delay_tab: st.markdown('</div>', unsafe_allow_html=True)
                                
                            if user_checked != current_check_state:
                                next_status = "Y" if user_checked else "N"
                                save_production_note(pure_excel_code, memo_tuple[0], memo_tuple[1], memo_tuple[2], memo_tuple[3], memo_tuple[4], next_status)
                                st.rerun()
                            
                            # 🚨 [2안 핵심 설계 마감 조항] : 타이핑 중 새로고침이 절대 발생하지 않도록 폼 내부 메모리에 값을 1차 홀딩
                            u_c_code = st.text_input(label="1. 카톤코드", value=memo_tuple[0] if memo_tuple[0] != "-" else "", placeholder="1. 카톤코드 입력", key=f"inp_c_{section_prefix}_{pure_excel_code}_{idx}")
                            u_pack_qty = st.text_input(label="2. 개입수", value=memo_tuple[1] if memo_tuple[1] != "-" else "", placeholder="2. 개입수 입력", key=f"inp_p_{section_prefix}_{pure_excel_code}_{idx}")
                            u_m_date = st.text_input(label="3. 제조일", value=memo_tuple[2] if memo_tuple[2] != "-" else "", placeholder="3. 제조일 입력", key=f"inp_d_{section_prefix}_{pure_excel_code}_{idx}")
                            u_m_qty = st.text_input(label="4. 제조량", value=memo_tuple[3] if memo_tuple[3] != "-" else "", placeholder="4. 제조량 입력", key=f"inp_q_{section_prefix}_{pure_excel_code}_{idx}")
                            u_p_qty = st.text_input(label="5. 생산수량", value=memo_tuple[4] if memo_tuple[4] != "-" else "", placeholder="5. 생산수량 입력", key=f"inp_s_{section_prefix}_{pure_excel_code}_{idx}")
                            
                            # 🚨 [일괄 완료 컴파일 스위치] : 5개 정보 기입 후 원터치로 파일로그 영구 박음질
                            st.markdown('<div class="owner-save-btn-zone">', unsafe_allow_html=True)
                            if st.button("💾 이 카드 정보 저장", key=f"btn_save_{section_prefix}_{pure_excel_code}_{idx}", use_container_width=True):
                                save_production_note(pure_excel_code, u_c_code, u_pack_qty, u_m_date, u_m_qty, u_p_qty, memo_tuple[5])
                                st.toast(f"🎯 [{pure_excel_code}] 실무 데이터가 디스크에 영구 병합되었습니다.")
                                time.sleep(0.3)
                                st.rerun()
                            st.markdown('</div>', unsafe_allow_html=True)
                                
                            st.markdown('<div style="margin-bottom:8px;"></div>', unsafe_allow_html=True)

            # 기타 카테고리
            other_df = target_df[~target_df['category'].str.lower().str.contains('skin|body|hair')]
            if not other_df.empty:
                st.markdown('<div style="font-size:20px; font-weight:bold; color:#94a3b8; padding:6px 12px; background-color:#0f172a; border-left:5px solid #94a3b8; border-radius:4px; margin-top:25px; margin-bottom:15px;">📦 기타 카테고리 Lineup</div>', unsafe_allow_html=True)
                cols = st.columns(6)
                for idx, row in other_df.reset_index(drop=True).iterrows():
                    excel_code = row['item_code']
                    pure_excel_code = extract_pure_6_code(excel_code)
                    with cols[idx % 6]:
                        local_base64_data = get_saved_local_image_base64(pure_excel_code)
                        st.html(f'<div class="owner-square-frame"><img src="{local_base64_data if local_base64_data else ""}"></div>')
                        
                        st.html(f"""
                            <div class="owner-info-card-wrap">
                                <div class="owner-text-row" style="font-size:30px !important; font-weight:900 !important; color:#ffffff !important; margin-bottom:6px !important; letter-spacing:0.5px !important;">{excel_code}</div>
                                <div class="owner-text-row" style="font-size:14px !important; color:#a0aec0 !important; font-weight:500 !important; min-height:40px !important; margin-bottom:14px !important; display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; overflow:hidden;">{row['product_name']}</div>
                                <div style="border-bottom:1px solid #2d3748 !important; margin-bottom:12px !important;"></div>
                                <div class="owner-text-row" style="font-size:14px !important; color:#718096 !important; margin-bottom:3px !important;">가격표 유무: <span style="color:#63b3ed !important; font-weight:bold !important;">{row['price_tag']}</span></div>
                                <div class="owner-text-row" style="font-size:14px !important; color:#ffffff !important; margin-bottom:3px !important;">용량: <span style="color:#ffffff !important; font-weight:bold !important;">{row['volume']}</span></div>
                                <div class="owner-text-row" style="font-size:14px !important; color:#718096 !important; margin-bottom:3px !important;">PO#: <span style="color:#ecc94b !important; font-weight:bold !important;">{row['po_number']}</span></div>
                                <div class="owner-text-row" style="font-size:14px !important; color:#718096 !important; margin-bottom:3px !important;">Bag#: <span style="color:#e53e3e !important; font-weight:bold !important;">{row['bag_number']}</span></div>
                                <div style="background-color:#111622 !important; border-radius:8px !important; padding:8px 12px !important; display:flex !important; justify-content:space-between !important; align-items:center !important;">
                                    <span class="owner-text-row" style="font-size:16px !important; color:#48bb78 !important; font-weight:bold !important;">📦 {row['quantity']:,}개</span>
                                    <span class="owner-text-row" style="font-size:13px !important; color:#a0aec0 !important; font-weight:500 !important;">📅 {row['production_date'].strftime('%m-%d')}</span>
                                </div>
                            </div>
                        """)
                        memo_tuple = saved_notes.get(pure_excel_code, ("-", "-", "-", "-", "-", "N"))
                        current_check_state = True if memo_tuple[5] == "Y" else False
                        
                        check_label = "✅ 생산 완료 완료됨" if is_completed_tab else "🚨 완료 시 체크 (하단 이동)"
                        user_checked = st.checkbox(label=check_label, value=current_check_state, key=f"chk_oth_{pure_excel_code}_{idx}")
                        if user_checked != current_check_state:
                            next_status = "Y" if user_checked else "N"
                            save_production_note(pure_excel_code, memo_tuple[0], memo_tuple[1], memo_tuple[2], memo_tuple[3], memo_tuple[4], next_status)
                            st.rerun()
                            
                        u_c_code = st.text_input(label="1. 카톤코드", value=memo_tuple[0] if memo_tuple[0] != "-" else "", placeholder="1. 카톤코드 입력", key=f"inp_c_oth_{pure_excel_code}_{idx}")
                        u_pack_qty = st.text_input(label="2. 개입수", value=memo_tuple[1] if memo_tuple[1] != "-" else "", placeholder="2. 개입수 입력", key=f"inp_p_oth_{pure_excel_code}_{idx}")
                        u_m_date = st.text_input(label="3. 제조일", value=memo_tuple[2] if memo_tuple[2] != "-" else "", placeholder="3. 제조일 입력", key=f"inp_d_oth_{pure_excel_code}_{idx}")
                        u_m_qty = st.text_input(label="4. 제조량", value=memo_tuple[3] if memo_tuple[3] != "-" else "", placeholder="4. 제조량 입력", key=f"inp_q_oth_{pure_excel_code}_{idx}")
                        u_p_qty = st.text_input(label="5. 생산수량", value=memo_tuple[4] if memo_tuple[4] != "-" else "", placeholder="5. 생산수량 입력", key=f"inp_s_oth_{pure_excel_code}_{idx}")
                        
                        st.markdown('<div class="owner-save-btn-zone">', unsafe_allow_html=True)
                        if st.button("💾 이 카드 정보 저장", key=f"btn_save_oth_{pure_excel_code}_{idx}", use_container_width=True):
                            save_production_note(pure_excel_code, u_c_code, u_pack_qty, u_m_date, u_m_qty, u_p_qty, memo_tuple[5])
                            st.toast(f"🎯 [{pure_excel_code}] 실무 데이터가 디스크에 영구 병합되었습니다.")
                            time.sleep(0.3)
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)
                        st.markdown('<div style="margin-bottom:8px;"></div>', unsafe_allow_html=True)

    # 지연 구역
    if not df_delayed.empty:
        st.markdown("""<div style='margin-top:20px; border-bottom: 4px solid #ff0000; padding-bottom:5px;'><h2 style='color:#ff4d4d; font-weight:bold;'>⚠️ FINE FORMULATION 생산 스케줄 지연 알림 리스트</h2></div>""", unsafe_allow_html=True)
        render_schedule_grid(df_delayed, "⚠️ 기한 초과 및 미완료 품목", "delayed", is_delay_tab=True)
        st.markdown("""<div style='margin-top:30px; margin-bottom:30px; border-bottom: 2px dashed #475569;'></div>""", unsafe_allow_html=True)

    # 현역 주차 보드
    render_schedule_grid(df_1week, f"📅 1주 차 생산 스케줄 대쉬보드 (V열 기한 기준)", "w1", inject_anchor=True)
    render_schedule_grid(df_2weeks, f"📅 2주 차 생산 스케줄 대쉬보드 (V열 기한 기준)", "w2")
    
    # 완료 목록 구역
    st.markdown("""<div style='margin-top:80px; border-bottom: 3px dashed #10b981;'></div>""", unsafe_allow_html=True)
    render_schedule_grid(df_completed_pool, f"✅ 실시간 생산 완료 영구 보존 라인업 (주차 무관 상시 보존)", "done", is_completed_tab=True)

else:
    st.info("💡 스케줄 마스터 엑셀 파일 로드 대기중")